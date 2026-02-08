from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .config import SETTINGS
from bot.shared.translator import translate_to_english

log = logging.getLogger("search")
PIN_RE = re.compile(r"\b\d{6}\b")


def _normalize_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.strip().lower())


def _best_column(columns: Iterable[str], candidates: list[str]) -> str | None:
    normalized = {col: _normalize_col(col) for col in columns}
    for cand in candidates:
        for col, norm in normalized.items():
            if cand in norm:
                return col
    return None


def _best_phone_column(rows: list[dict[str, Any]]) -> str | None:
    if not rows:
        return None
    columns = list(rows[0].keys())
    tokens = ["phone", "mobile", "tel", "contact", "call", "number", "whatsapp"]
    candidates = []
    for col in columns:
        norm = _normalize_col(col)
        if any(token in norm for token in tokens):
            candidates.append(col)

    if not candidates:
        candidates = columns

    best_col = None
    best_score = 0
    for col in candidates:
        score = 0
        for row in rows:
            value = str(row.get(col, "") or "")
            digits = re.sub(r"\D+", "", value)
            if len(digits) >= 6:
                score += 10 + min(len(digits), 15)
        if score > best_score:
            best_score = score
            best_col = col

    return best_col if best_score > 0 else None


def _load_excel(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        log.warning("excel_not_found", extra={"path": str(path)})
        return []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
    except Exception:
        log.exception("excel_read_failed", extra={"path": str(path)})
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    results: list[dict[str, Any]] = []
    for row in rows[1:]:
        record: dict[str, Any] = {}
        for idx, value in enumerate(row):
            key = headers[idx] if idx < len(headers) else f"col_{idx}"
            record[key] = "" if value is None else value
        results.append(record)
    return results


def _standardize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    columns = list(rows[0].keys())
    col_name = _best_column(columns, ["name", "hospital", "garage", "center"])
    col_city = _best_column(columns, ["city", "town"])
    col_state = _best_column(columns, ["state"])
    col_pin = _best_column(columns, ["pincode", "pin", "zipcode", "postal"])
    col_address = _best_column(columns, ["address", "addr", "location"])
    col_phone = _best_phone_column(rows) or _best_column(columns, ["phone", "contact", "mobile", "tel"])

    standardized = []
    for row in rows:
        standardized.append(
            {
                "name": str(row.get(col_name, "")).strip() if col_name else "",
                "city": str(row.get(col_city, "")).strip() if col_city else "",
                "state": str(row.get(col_state, "")).strip() if col_state else "",
                "pincode": str(row.get(col_pin, "")).strip() if col_pin else "",
                "address": str(row.get(col_address, "")).strip() if col_address else "",
                "phone": str(row.get(col_phone, "")).strip() if col_phone else "",
                "raw": row,
            }
        )
    return standardized


_HOSPITAL_CACHE: list[dict[str, Any]] | None = None
_GARAGE_CACHE: list[dict[str, Any]] | None = None


def load_hospitals() -> list[dict[str, Any]]:
    global _HOSPITAL_CACHE
    if _HOSPITAL_CACHE is None:
        rows = _load_excel(SETTINGS.data_paths.hospitals_xlsx)
        _HOSPITAL_CACHE = _standardize_rows(rows)
    return _HOSPITAL_CACHE


def load_garages() -> list[dict[str, Any]]:
    global _GARAGE_CACHE
    if _GARAGE_CACHE is None:
        rows = _load_excel(SETTINGS.data_paths.garages_xlsx)
        _GARAGE_CACHE = _standardize_rows(rows)
    return _GARAGE_CACHE


def _score_row(row: dict[str, Any], tokens: list[str], pincode: str | None) -> int:
    score = 0
    haystack = " ".join(
        [
            row.get("name", ""),
            row.get("city", ""),
            row.get("state", ""),
            row.get("address", ""),
            row.get("pincode", ""),
        ]
    ).lower()
    for token in tokens:
        if token and token in haystack:
            score += 2
    if pincode and row.get("pincode") and pincode in str(row.get("pincode")):
        score += 5
    return score


def _search(rows: list[dict[str, Any]], query: str) -> list[dict[str, Any]]:
    if not rows:
        return []
    tokens = [t for t in re.split(r"\s+", query.lower()) if t]
    pincode_match = PIN_RE.search(query)
    pincode = pincode_match.group(0) if pincode_match else None

    scored = []
    for row in rows:
        score = _score_row(row, tokens, pincode)
        if score > 0:
            scored.append((score, row))
    scored.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in scored]


def search_hospitals(query: str, limit: int = 3) -> list[dict[str, Any]]:
    rows = load_hospitals()
    search_query = translate_to_english(query)
    results = _search(rows, search_query)
    return results[:limit]


def search_garages(query: str, limit: int = 1) -> list[dict[str, Any]]:
    rows = load_garages()
    search_query = translate_to_english(query)
    results = _search(rows, search_query)
    return results[:limit]


def extract_pincode(text: str) -> str | None:
    match = PIN_RE.search(text)
    return match.group(0) if match else None
